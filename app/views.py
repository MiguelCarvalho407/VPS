from datetime import timedelta, datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .forms import *
from .models import *
from django.contrib import messages
from django.template.loader import render_to_string
from django.contrib.sites.shortcuts import get_current_site
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import EmailMessage, send_mail
# from .tokens import account_activation_token
from django.contrib.auth import get_user_model
import json
from django.urls import reverse
from django.utils import timezone
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.db.models import Q
from calendar import monthrange
from math import ceil
import openpyxl
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.contrib.auth import update_session_auth_hash


# PÁGINA DE ERRO 404 #
def custom_404(request, exception):
    return render(request, 'ERRORS/404.html', status=404)


# def base(request):
#     return render(request, 'base.html')


def signup(request):
    if request.method == 'POST':
        form = CriarContaForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']

            if Utilizadores.objects.filter(email=email).exists():
                form.add_error('email', 'Este email já está em uso.')
            else:
                utilizador = Utilizadores.objects.create_user(
                    username=form.cleaned_data['nome'],
                    email=form.cleaned_data['email'],
                    password=form.cleaned_data['password'],
                    genero=form.cleaned_data['genero'],
                    data_nascimento=form.cleaned_data['data_nascimento'],
                    contacto=form.cleaned_data['contacto'],
                    nif=form.cleaned_data['nif'],
                    pretende_recibo=form.cleaned_data['pretende_recibo'],
                    profissao=form.cleaned_data['profissao'],
                    classificacao_esforco_na_profissao=form.cleaned_data['classificacao_esforco_na_profissao'],
                    fumador=form.cleaned_data['fumador'],
                    problemas_saude=form.cleaned_data['problemas_saude'],
                    limitacoes_para_pratica_exercicio_fisico=form.cleaned_data['limitacoes_para_pratica_exercicio_fisico'],
                    como_teve_conhecimento_existencia_fitclub=form.cleaned_data['como_teve_conhecimento_existencia_fitclub'],
                    altura=form.cleaned_data['altura'],
                    codigo_postal=form.cleaned_data['codigo_postal'],
                    localidade=form.cleaned_data['localidade'],
                )
                utilizador.save()

                return redirect('login')
    else:
        form = CriarContaForm()

    return render(request, 'CONTAS/criar_conta.html', {'form': form})


def dologin(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('email')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=email, password=password)
            if user is not None:
                login(request, user)
                return redirect('fcBase')
            else:
                form.add_error(None, 'Email ou password incorretos!')
    else:
        form = LoginForm()
    return render(request, 'CONTAS/login.html', {'form': form})



@login_required
def alterar_senha(request):
    if request.method == "POST":
        form = AlterarSenhaForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Mantém o usuário logado após a mudança de senha
            return redirect('alterar_senha_concluido')
    else:
        form = AlterarSenhaForm(user=request.user)
    
    return render(request, 'alterar_senha.html', {'form': form})

@login_required
def alterar_senha_concluido(request):
    return render(request, 'alterar_senha_concluido.html')




def logout_view(request):
    logout(request)
    return redirect('login')


# def activateEmail(request, user, to_email):
#     mail_subject = "FitClub - Confirmação de email."
#     message = render_to_string("template_activate_account.html", {
#         'user': user,
#         'domain': get_current_site(request).domain,
#         'uid': urlsafe_base64_encode(force_bytes(user.pk)),
#         'token': account_activation_token.make_token(user),
#         "protocol": 'https' if request.is_secure() else 'http'
#     })
#     email = EmailMessage(mail_subject, message, to=[to_email])
#     if email.send():
#         messages.success(request, f'Email enviado para <b>{to_email}</b> clica no link enviado para confirmar e completar o registo \
#                 <b>Nota:</b> Verifica no spam. <b>ATENÇÃO:</b>  O link expira em 1 hora')
#     else:
#         messages.error(request, f'Problema ao enviar email para {to_email}, confirma que inseriste o email corretamente.')


# def activate(request, uidb64, token):
#     User = get_user_model()
#     try:
#         uid = force_str(urlsafe_base64_decode(uidb64))
#         user = User.objects.get(pk=uid)
#     except:
#         user = None

#     if user is not None and account_activation_token.check_token(user, token):
#         user.is_active = True
#         user.save()

#         messages.success(request, "Conta confirmada! Faz login para teres acesso.")
#         return redirect('login')
#     else:
#         messages.error(request, "Link de confirmação inválido!")

#     return redirect('signup')





########## FITCLUB APP ##########

def acesso_negado(request):
    return render(request, 'ERRORS/acesso_negado.html')


REGIME_PRECOS = {
    'regime_livre': '35€',
    'regime_livre+plano_alimentar': '40€',
    'regime_basico': '30€',
    'regime_basico+plano_alimentar': '37,5€',
    'so_treinos_online': '10€',
    'pack_6_aulas': '25€',
    'pack_4_aulas': '20€',
    'kids': '20€',
}

MESES_PT = [
    'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
]

def mes_anterior(ref_date=None):
    if ref_date is None:
        ref_date = date.today()

    if ref_date.month == 1:
        return ref_date.year - 1, 12
    return ref_date.year, ref_date.month - 1

def is_trimestral(mes):
    return mes in [3, 6, 9, 12]

@login_required
def escolher_regime(request):
    hoje = date.today()
    ano, mes = mes_anterior(hoje)

    trimestre = is_trimestral(mes)

    if request.method == 'POST':
        form = EscolherRegimeForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                regime = form.cleaned_data['regime']
                request.user.regime = regime
                request.user.regime_data = date.today()
                request.user.save()

                ano, mes = mes_anterior(request.user.regime_data)

                AvaliacaoMensal.objects.get_or_create(
                    utilizador=request.user,
                    ano=ano,
                    mes=mes,
                    nivel_satisfacao=form.cleaned_data['nivel_satisfacao']
                )

                if trimestre:
                    nivel_satisfacao_trimestral = form.cleaned_data.get('nivel_satisfacao_trimestral')
                    AvaliacaoTrimestral.objects.get_or_create(
                        utilizador=request.user,
                        ano=ano,
                        mes=mes,
                        nivel_satisfacao=nivel_satisfacao_trimestral
                    )

            hoje = request.user.regime_data
            mes_nome = MESES_PT[hoje.month - 1]
            ano = hoje.year
            preco = REGIME_PRECOS.get(regime, '')

            regime_display = request.user.get_regime_display()

            message = (
                f'Olá {request.user.username},\n\n'
                f'De acordo com a sua escolha, confirmamos que o regime escolhido para o mês de {mes_nome} de {ano} foi o "{regime_display}". \n'
                f'Deverá efetuar o pagamento de {preco} até ao dia 8 deste mês.\n\n'
                f'FITCLUB - Mantém-te FIT! Mantém-te Forte!\n'
                'Email gerado automaticamente. Qualquer questão entre em contacto no WhatsApp.'
            )

            send_mail(
                subject='FITCLUB - Regime',
                message=message,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[request.user.email],
                fail_silently=False,
            )

            return redirect('fcBase')
    else:
        form = EscolherRegimeForm()
    return render(request, 'escolher_regime.html', {'form': form, 'trimestre': trimestre})




@login_required
def fcBase(request):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')

    utilizador = request.user

    # OBTER O MÊS E ANO ATUAL
    ano = int(request.GET.get('ano', datetime.now().year))
    mes = int(request.GET.get('mes', datetime.now().month))

    reservas = Reservas.objects.filter(
        utilizador=request.user,
        treino__data_inicio__year=ano,
        treino__data_inicio__month=mes,
    ).select_related('treino')

    # RESERVAS CONFIRMADAS
    reservas_confirmadas = reservas.filter(confirmado=True).count()

    # # RESERVAS NÃO CONFIRMADOS
    # total_reservas_nao_confirmadas = reservas.filter(confirmado=False).count()

    # # CONTAR O TOTAL DE TREINOS RESERVADOS
    # total_reservas = reservas.count()

    # Verifica ausências pendentes
    ausencias_pendentes = Ausencias.objects.filter(
        utilizador=utilizador,
        taxa_respondida='nao',
        reserva__confirmado=False
    ).select_related('reserva__treino')

    if request.method == 'POST':
        ausencia_id = request.POST.get('ausencia_id')
        resposta = request.POST.get('resposta')

        if ausencia_id and resposta in ['sim', 'nao']:
            ausencia = get_object_or_404(Ausencias, id=ausencia_id, utilizador=utilizador)
            ausencia.taxa_respondida = resposta
            ausencia.taxa_pago = True if resposta == 'sim' else False
            ausencia.save()
            return redirect('fcBase')

    return render(request, 'FC_APP/fcBase.html', {'ausencias': ausencias_pendentes, 'reservas': reservas, 'reservas_confirmadas': reservas_confirmadas})



@login_required
def criar_tipo_treino(request):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')
    
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    if request.method == 'POST':
        form = CriarTipoTreinoForm(request.POST or None)
        if form.is_valid():
            form.save()

            return redirect('criar_treino')
        
    else:
        form = CriarTipoTreinoForm()


    return render(request, 'FC_APP/fcCriar_tipo_treino.html', {'form': form})


@login_required
def criar_treinos(request):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    if request.method == "POST":
        form = CriarTreinoForm(request.POST)
        if form.is_valid():
            tipo_treino = form.cleaned_data['tipo_treino']
            tipo_treino_nome = form.cleaned_data['tipo_treino_nome']
            data_inicio = form.cleaned_data['data_inicio']
            data_fim = form.cleaned_data['data_fim']
            hora_inicio = form.cleaned_data['hora_inicio']
            hora_fim = form.cleaned_data['hora_fim']
            max_participantes = form.cleaned_data['max_participantes']
            max_lista_espera = form.cleaned_data['max_lista_espera']
            dias_da_semana = form.cleaned_data['dia_da_semana']
            reservas_horas_antes = form.cleaned_data['reservas_horas_antes']
            reservas_horas_fecho = form.cleaned_data['reservas_horas_fecho']

            # Mapeamento para exibição legível do tipo de treino
            TIPOS_TREINO_DISPLAY = dict(Treino.TIPO_TREINO_CHOICES)
            tipo_treino_display = TIPOS_TREINO_DISPLAY.get(tipo_treino, tipo_treino)  # Fallback para segurança

            current_date = data_inicio
            while current_date <= data_fim:
                for dia_da_semana in dias_da_semana:
                    dia_semana_map = {
                        'segunda-feira': 0,
                        'terça-feira': 1,
                        'quarta-feira': 2,
                        'quinta-feira': 3,
                        'sexta-feira': 4,
                        'sábado': 5,
                        'domingo': 6,
                    }

                    if current_date.weekday() == dia_semana_map[dia_da_semana]:
                        # Verifica se já existe um treino com o mesmo tipo e horário
                        if Treino.objects.filter(
                            tipo_treino=tipo_treino,
                            data_inicio=current_date,
                            hora_inicio=hora_inicio
                        ).exists():
                            messages.error(
                                request, 
                                f'Já existe um treino "{tipo_treino_display}" às {hora_inicio} dia {current_date.strftime("%d/%m/%Y")}.'
                            )
                            return redirect('criar_treino')

                        # Criação do treino
                        Treino.objects.create(
                            tipo_treino=tipo_treino,
                            tipo_treino_nome=tipo_treino_nome,
                            data_inicio=current_date,
                            data_fim=current_date,
                            hora_inicio=hora_inicio,
                            hora_fim=hora_fim,
                            max_participantes=max_participantes,
                            max_lista_espera=max_lista_espera,
                            dia_da_semana=dia_da_semana,
                            reservas_horas_antes=reservas_horas_antes,
                            reservas_horas_fecho=reservas_horas_fecho,
                        )
                current_date += timedelta(days=1)

            return redirect('calendario')

    else:
        form = CriarTreinoForm()

    return render(request, 'FC_APP/fcCriar_treino.html', {'form': form})



@login_required
def definicoes(request):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    regime = request.user.get_regime_display()

    if request.method == 'POST':
        form = InformacoesPessoaisForm(request.POST, request.FILES, instance=request.user)

        if form.is_valid():
            form.save()
            return redirect('calendario')
    else:
        form = InformacoesPessoaisForm(instance=request.user)

    return render(request, 'FC_APP/fcDefinicoes.html', {'form': form, 'regime': regime})



@login_required
def calendario(request):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')

    year = int(request.GET.get("year", timezone.now().year))
    month = int(request.GET.get("month", timezone.now().month))

    treinos = Treino.objects.filter(
        data_inicio__year=year,
        data_inicio__month=month
    ).order_by('hora_inicio')

    current_datetime = timezone.now()

    treinos_data = []
    for treino in treinos:
        user_reserved = Reservas.objects.filter(utilizador=request.user, treino=treino).exists()
        treino_inicio = timezone.make_aware(
            timezone.datetime.combine(treino.data_inicio, treino.hora_inicio)
        )
        has_passed = current_datetime > treino_inicio

        treinos_data.append({
            'id': treino.id,
            'tipo_treino': treino.get_tipo_treino_display(),
            'tipo_treino_nome': treino.tipo_treino_nome.nome if treino.tipo_treino_nome else None,
            'data_inicio': treino.data_inicio.strftime('%Y-%m-%d'),
            'hora_inicio': treino.hora_inicio.strftime('%H:%M'),
            'hora_fim': treino.hora_fim.strftime('%H:%M'),
            'max_participantes': treino.max_participantes,
            'reservados': treino.total_reservas(),
            'reservar_url': reverse('reservas', args=[treino.id]),
            'detalhes_url': reverse('reservas_detalhes', args=[treino.id]),
            'user_reserved': user_reserved,
            'has_passed': has_passed,
        })

    # Se for AJAX (fetch), retorna JSON
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse(treinos_data, safe=False)

    # Se não, renderiza a página inicial
    return render(request, 'FC_APP/fcCalendario.html', {
        "treinos_json": json.dumps(treinos_data),
        "year": year,
        "month": month,
    })



@login_required
def dadosbiometricos(request):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')

    ano = int(request.GET.get('ano', datetime.now().year))
    mes = int(request.GET.get('mes', datetime.now().month))

    altura = request.user.altura

    dados_biometricos = Dados_biometricos.objects.filter(
        utilizador=request.user,
        data_registo__year=ano,
        data_registo__month=mes,
    ).first()

    # Calcular idade do utilizador
    if request.user.data_nascimento:
        hoje = date.today()
        idade = hoje.year - request.user.data_nascimento.year - (
            (hoje.month, hoje.day) < (request.user.data_nascimento.month, request.user.data_nascimento.day)
        )
    else:
        idade = "Não informado"

    if not dados_biometricos:
        dados_biometricos = Dados_biometricos(utilizador=request.user, data_registo=datetime(ano, mes, 1))

    if request.method == 'POST':
        form = DadosBiometricosForm(request.POST, instance=dados_biometricos)
        if form.is_valid():
            novo_dado = form.save(commit=False)
            novo_dado.utilizador = request.user
            # novo_dado.data_registo = datetime(ano, mes, 1)
            novo_dado.save()
            return redirect('dadosbiometricos')
        else:
            print(form.errors)  # Verifica se há erros no formulário

    else:
        form = DadosBiometricosForm(instance=dados_biometricos)

    meses = range(1, 13)

    return render(
        request,
        'FC_APP/fcDados_biometricos.html',
        {
            'form': form,
            'dados_biometricos': dados_biometricos,
            'ano': ano,
            'mes': mes,
            'meses': meses,
            'idade': idade,
            'altura': altura,
        }
    )



@login_required
def editardadosbiometricos(request, user_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    # Obtém o mês e o ano da query string ou usa os valores atuais
    ano = int(request.GET.get('ano', datetime.now().year))
    mes = int(request.GET.get('mes', datetime.now().month))

    # VAI BUSCAR O UTILIZADOR PARA NO HTML APARECER O NOME DE UTILIZADOR QUE ESTOU A EDITAR
    utilizador = Utilizadores.objects.get(id=user_id)

    #IR BUSCAR A ALTURA À OUTRA TABELA
    altura = utilizador.altura

    # Filtra os dados biométricos com base no mês, ano e utilizador
    dadobiometrico = Dados_biometricos.objects.filter(
        utilizador_id=user_id, 
        data_registo__year=ano,
        data_registo__month=mes,
    ).first()

    # Calcular idade do utilizador
    if utilizador.data_nascimento:
        hoje = date.today()
        idade = hoje.year - utilizador.data_nascimento.year - (
            (hoje.month, hoje.day) < (utilizador.data_nascimento.month, utilizador.data_nascimento.day)
        )
    else:
        idade = "Não informado"

    # Se nenhum dado for encontrado, cria um novo objeto para edição
    if not dadobiometrico:
        dadobiometrico = Dados_biometricos(utilizador_id=user_id, data_registo=datetime(ano, mes, 1))

    if request.method == 'POST':
        form = EditarDadosBiometricos(request.POST, instance=dadobiometrico)
        if form.is_valid():
            novo_dado = form.save(commit=False)
            novo_dado.utilizador_id = user_id
            novo_dado.data_registo = datetime(ano, mes, 1)
            novo_dado.save()
            return redirect(
                f'{reverse("editardadosbiometricos", kwargs={"user_id": user_id})}?ano={ano}&mes={mes}'
            )
    else:
        form = EditarDadosBiometricos(instance=dadobiometrico)

    # Lista de meses para o dropdown
    meses = range(1, 13)

    # Carrega as reservas do usuário (usando `utilizador_id` como filtro)
    reservas = Reservas.objects.filter(
        utilizador_id=user_id,
        treino__data_inicio__year=ano,
        treino__data_inicio__month=mes,
    ).select_related('treino')

    # Conta o número de reservas confirmadas
    total_reservas_confirmadas = reservas.filter(confirmado=True).count()

    # Conta o número de reservas não confirmadas
    total_reservas_nao_confirmadas = reservas.filter(confirmado=False).count()

    # Conta o número total de reservas
    total_reservas = reservas.count()

    # Cálculo da média semanal de frequência
    _, total_dias_no_mes = monthrange(ano, mes)
    semanas_uteis_no_mes = ceil(total_dias_no_mes / 7)
    dias_uteis_totais = semanas_uteis_no_mes * 5
    media_semanal = (total_reservas_confirmadas / dias_uteis_totais) * 100 if dias_uteis_totais > 0 else 0

    return render(
        request,
        'FC_APP/fcEditar_dados_biometricos.html',
        {
            'form': form,
            'utilizador': utilizador,
            'dadobiometrico': dadobiometrico,
            'ano': ano,
            'mes': mes,
            'meses': meses,
            'total_reservas_confirmadas': total_reservas_confirmadas,
            'total_reservas_nao_confirmadas': total_reservas_nao_confirmadas,
            'total_reservas': total_reservas,
            'media_semanal': media_semanal,
            'reservas': reservas,
            'assiduidade_url': reverse('verassiduidade', kwargs={'user_id': user_id}),
            'idade': idade,
            'altura': altura,
        },
    )

@login_required
def detalhe_dadosbiometricos(request, user_id):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    dadobiometrico = get_object_or_404(Dados_biometricos, user__id=user_id)
    return render(request, 'FC_APP/fcDetalheUtilizadores.html', {'dadobiometrico': dadobiometrico})




@login_required
def membros(request):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    ano_atual = date.today().year
    anos = list(range(ano_atual - 2, ano_atual + 1))
    meses = list(range(1, 13))

    query = request.GET.get('q', '')  # Obtém o termo de pesquisa da URL
    utilizadores = Utilizadores.objects.all()

    if query:
        utilizadores = utilizadores.filter(username__icontains=query)  # Filtra pelo nome de usuário

    # Paginação
    paginator = Paginator(utilizadores, 10)  # 10 utilizadores por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'FC_APP/fcMembros.html', {'page_obj': page_obj, 'query': query, 'anos': anos, 'meses': meses})


@login_required
def editar_utilizador(request, user_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    utilizador = get_object_or_404(Utilizadores, id=user_id)

    if request.method == 'POST':
        form = EditarInformacoesPessoaisForm(request.POST, instance=utilizador)
        if form.is_valid():
            form.save()
            return redirect('detalhe_utilizador', user_id=user_id)
    else:
        form = EditarInformacoesPessoaisForm(instance=utilizador)

    return render(request, 'FC_APP/fcEditarUtilizador.html', {'form': form, 'utilizador': utilizador})


@login_required
def apagarconta(request, user_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')
    
    user = get_object_or_404(Utilizadores, id=user_id)
    
    if request.user == user:
        return redirect('membros')

    user.delete()
    return redirect('membros')  # Redireciona para a lista de membros



@login_required
def detalhe_utilizador(request, user_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    perfil = get_object_or_404(Utilizadores, id=user_id)
    return render(request, 'FC_APP/fcDetalheUtilizadores.html', {'perfil': perfil})


######################################################################



# @login_required
# def reservas(request, treino_id):
#     treino = get_object_or_404(Treino, id=treino_id)

#     # Verificar se o treino já começou
#     treino_inicio = timezone.make_aware(
#         timezone.datetime.combine(treino.data_inicio, treino.hora_inicio)
#     )
#     if timezone.now() > treino_inicio:
#         return render(request, 'FC_RESERVAS/fcReservas_tempo_passado.html', {'treino': treino})

#     # Verificar se o utilizador já tem reserva
#     reserva = Reservas.objects.filter(utilizador=request.user, treino=treino).first()
#     if reserva:
#         # Cancelar reserva e mover o primeiro da lista de espera para as reservas
#         reserva.delete()
#         primeiro_na_lista = ListaEspera.objects.filter(treino=treino).order_by('data_entrada').first()
#         if primeiro_na_lista:
#             Reservas.objects.create(utilizador=primeiro_na_lista.utilizador, treino=treino)
#             primeiro_na_lista.delete()
#         return redirect('calendario')

#     # Verificar se o treino está cheio
#     if Reservas.objects.filter(treino=treino).count() >= treino.max_participantes:
#         # Adicionar o utilizador à lista de espera
#         ListaEspera.objects.get_or_create(utilizador=request.user, treino=treino)

#         # Obter a lista de espera para renderizar no template
#         lista_espera = ListaEspera.objects.filter(treino=treino).order_by('data_entrada')
#         reservas_confirmadas = Reservas.objects.filter(treino=treino)
        
#         return render(request, 'FC_RESERVAS/fcReservas_detalhes.html', {
#             'treino': treino,
#             'lista_espera': lista_espera,
#             'reservas': reservas_confirmadas,  # Também passar reservas confirmadas
#         })

#     # Criar uma nova reserva
#     Reservas.objects.create(utilizador=request.user, treino=treino)
#     return redirect('calendario')

#     # CRIAR UMA NOVA RESERVA
#     Reservas.objects.create(utilizador=request.user, treino=treino)
#     return redirect('calendario')





from django.db import transaction

@login_required
def reservas(request, treino_id):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    with transaction.atomic():
        treino = (Treino.objects.select_for_update().get(id=treino_id))


        if not treino.reservas_abertas():
            now = timezone.now()
            inicio_treino = timezone.make_aware(
                timezone.datetime.combine(treino.data_inicio, treino.hora_inicio)
            )
            abertura_reservas = inicio_treino - timedelta(hours=treino.reservas_horas_antes)
            fechamento_reservas = inicio_treino - timedelta(hours=treino.reservas_horas_fecho)

            if now < abertura_reservas:
                contexto = {'treino': treino, 'abertura_reservas': abertura_reservas}
                return render(request, 'FC_RESERVAS/fcReservas_reservas_nao_abertas.html', contexto)
            else:
                contexto = {'treino': treino, 'fechamento_reservas': fechamento_reservas}
                return render(request, 'FC_RESERVAS/fcReservas_reservas_fechadas.html', contexto)

        treino_inicio = timezone.make_aware(
            timezone.datetime.combine(treino.data_inicio, treino.hora_inicio)
        )
        if timezone.now() > treino_inicio:
            return render(request, 'FC_RESERVAS/fcReservas_tempo_passado.html', {'treino': treino})

        # VERIFICAR SE O UTILIZADOR JÁ TEM UMA RESERVA NO MESMO DIA
        reserva_existente = Reservas.objects.filter(
            utilizador=request.user,
            treino__data_inicio=treino.data_inicio  # Verifica se já tem treino no mesmo dia
        ).first()

        if reserva_existente:
            if reserva_existente.treino == treino:
                # SE O USUÁRIO ESTÁ TENTANDO CANCELAR A MESMA RESERVA, CANCELAMOS
                reserva_existente.delete()
                primeiro_na_lista = ListaEspera.objects.filter(treino=treino).order_by('data_entrada').first()
                if primeiro_na_lista:
                    Reservas.objects.create(utilizador=primeiro_na_lista.utilizador, treino=treino)
                    primeiro_na_lista.delete()
                return redirect('calendario')
            else:
                # SE JÁ TIVER OUTRO TREINO NO MESMO DIA, BLOQUEAMOS A NOVA RESERVA
                messages.error(request, "Não é permitido fazer mais do que uma reserva por dia.")
                return redirect('calendario')

        # 🚀 NOVA VERIFICAÇÃO: IMPEDIR RESERVA SE O UTILIZADOR ESTIVER NA LISTA DE ESPERA DE OUTRO TREINO
        espera_existente = ListaEspera.objects.filter(
            utilizador=request.user,
            treino__data_inicio=treino.data_inicio
        ).exclude(treino=treino).exists()  # Verifica se está em lista de espera de outro treino

        if espera_existente:
            messages.error(request, "Não podes reservar esta aula pois já estás na lista de espera de outra.")
            return redirect('calendario')

        # REGIMES
        if request.user.regime == 'regime_basico' or request.user.regime == 'regime_basico+plano_alimentar':
            hoje = treino.data_inicio
            inicio_semana = hoje - timedelta(days=hoje.weekday())
            fim_semana = inicio_semana + timedelta(days=6)

            reservas_semana = Reservas.objects.filter(
                utilizador=request.user,
                treino__data_inicio__range=(inicio_semana, fim_semana)
            ).count()

            if reservas_semana >= 2:
                messages.error(request, "Com o regime básico, só podes marcar 2 treinos por semana.")
                return redirect('calendario')


        if request.user.regime == 'so_treinos_online':
            messages.error(request, 'Com regime online não podes fazer reservas em treinos presenciais.')
            return redirect('calendario')


        if request.user.regime == 'pack_6_aulas':
            hoje = treino.data_inicio
            inicio_mes = hoje.replace(day=1)

            if hoje.month== 12:
                fim_mes = hoje.replace(year=hoje.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                fim_mes = hoje.replace(month=hoje.month + 1, day=1) - timedelta(days=1)

            reservas_mes = Reservas.objects.filter(
                utilizador=request.user,
                treino__data_inicio__range=(inicio_mes, fim_mes)
            ).count()

            if reservas_mes >= 6:
                messages.error(request, 'Com o regime de pack 6 aulas, só podes marcar 6 treinos por mês.')
                return redirect('calendario')


        if request.user.regime == 'pack_4_aulas':
            hoje = treino.data_inicio
            inicio_mes = hoje.replace(day=1)

            if hoje.month== 12:
                fim_mes = hoje.replace(year=hoje.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                fim_mes = hoje.replace(month=hoje.month + 1, day=1) - timedelta(days=1)

            reservas_mes = Reservas.objects.filter(
                utilizador=request.user,
                treino__data_inicio__range=(inicio_mes, fim_mes)
            ).count()

            if reservas_mes >= 4:
                messages.error(request, 'Com o regime de pack 4 aulas, só podes marcar 4 treinos por mês.')
                return redirect('calendario')


        if request.user.regime == 'regime_kids':
            tipo = 17

            if not treino.tipo_treino_nome or treino.tipo_treino_nome.id != tipo:
                messages.error(request, f' Com o "Regime Kids" só podes reservar os treinos "Pais e Filhos".')
                return redirect('calendario')

        # IMPEDIR A MARCAÇÃO QUANDO O REGIME NÃO É O 'regime_kids'
        # tipo_kids = 17
        # if treino.tipo_treino_nome and treino.tipo_treino_nome.id == tipo_kids:
        #     if request.user.regime != 'regime_kids':
        #         messages.error(request, f'Não podes reservar um treino do tipo "Pais e Filhos" com o regime atual.')
        #         return redirect('calendario')


        # VERIFICAR SE O TREINO ESTÁ CHEIO
        # if Reservas.objects.filter(treino=treino).count() >= treino.max_participantes:
        #     if ListaEspera.objects.filter(treino=treino).count() >= treino.max_lista_espera:
        #         return render(request, 'lista_espera_full.html', {'treino': treino})

        #     ListaEspera.objects.get_or_create(utilizador=request.user, treino=treino)
        #     return render(request, 'reservas_lista_espera.html', {'treino': treino})
        
        reservas_qs = Reservas.objects.select_for_update().filter(treino=treino)

        if reservas_qs.count() >= treino.max_participantes:
            lista_qs = ListaEspera.objects.select_for_update().filter(treino=treino)

            if lista_qs.count() >= treino.max_lista_espera:
                return render(request, 'lista_espera_full.html', {'treino': treino})
            
            ListaEspera.objects.get_or_create(utilizador=request.user, treino=treino)
            return render(request, 'reservas_lista_espera.html', {'treino': treino})

    # CRIAR RESERVA
        Reservas.objects.create(utilizador=request.user, treino=treino)

    return redirect('calendario')







@login_required
def adicionar_utilizador_treino(request, treino_id):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')

    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    treino = get_object_or_404(Treino, id=treino_id)

    # Contar o número de reservas confirmadas e de pessoas na lista de espera
    reservas_confirmadas = Reservas.objects.filter(treino=treino).count()
    pessoas_em_espera = ListaEspera.objects.filter(treino=treino).count()

    if request.method == 'POST':
        usuario_id = request.POST.get('usuario_id')
        usuario = get_object_or_404(Utilizadores, id=usuario_id)

        # Se houver vagas, adiciona o utilizador ao treino
        if reservas_confirmadas < treino.max_participantes:
            reserva, created = Reservas.objects.get_or_create(
                utilizador=usuario,
                treino=treino,
                defaults={'confirmado': None}  # Reserva confirmada diretamente
            )
            messages.success(request, f"{usuario.username} foi adicionado ao treino.")
        else:
            # Se não houver vagas, verifica se a lista de espera está cheia
            if pessoas_em_espera >= treino.max_lista_espera:
                messages.error(request, f"A lista de espera está cheia! {usuario.username} não pode ser adicionado.")
            else:
                ListaEspera.objects.get_or_create(utilizador=usuario, treino=treino)
                messages.info(request, f"O treino está cheio. {usuario.username} foi adicionado à lista de espera.")

        return redirect('reservas_detalhes', treino_id=treino.id)

    # Excluir usuários que já estão no treino ou na lista de espera
    utilizadores_com_reserva = Reservas.objects.filter(treino=treino).values_list('utilizador_id', flat=True)
    utilizadores_em_espera = ListaEspera.objects.filter(treino=treino).values_list('utilizador_id', flat=True)
    utilizadores_excluidos = list(utilizadores_com_reserva) + list(utilizadores_em_espera)

    utilizadores_disponiveis = Utilizadores.objects.exclude(id__in=utilizadores_excluidos)

    return render(request, 'adicionar_utilizador_treino.html', {'treino': treino, 'usuarios': utilizadores_disponiveis})












@login_required
def lista_espera_view(request, treino_id):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    treino = get_object_or_404(Treino, id=treino_id)
    lista_espera = ListaEspera.objects.filter(treino=treino).order_by('data_entrada')
    return render(request, 'FC_RESERVAS/fcLista_espera.html', {'lista_espera': lista_espera, 'treino': treino})



@login_required
def cancelar_lista_espera(request, treino_id):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    treino = get_object_or_404(Treino, id=treino_id)

    # Verificar se o usuário enviou um ID de espera válido
    espera_id = request.POST.get("espera_id")

    if espera_id:
        lista_espera = ListaEspera.objects.filter(id=espera_id, treino=treino).first()

        if lista_espera:
            # Permitir que is_staff remova qualquer um e que o próprio usuário saia da lista
            if request.user.is_staff or lista_espera.utilizador == request.user:
                lista_espera.delete()


    return redirect('reservas_detalhes', treino_id=treino.id)







################################################################################################



@login_required
def reservas_detalhes(request, treino_id):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')

    treino = get_object_or_404(Treino, id=treino_id)
    reservas = Reservas.objects.filter(treino=treino)
    lista_espera = ListaEspera.objects.filter(treino=treino).order_by('data_entrada')

    if request.method == 'POST':
        action = request.POST.get('action')

        #Marcar todos os presentes
        if action == 'confirmar_todos':
            reservas.update(confirmado=True)
            return redirect('reservas_detalhes', treino_id=treino.id)

        elif action in ['presente', 'ausente']:
            reservas_id = request.POST.get('reservas_id')
            reserva = get_object_or_404(Reservas, id=reservas_id)
            reserva.confirmado = (action == 'presente')
            reserva.save()

            if action == 'ausente':
                if not Ausencias.objects.filter(reserva=reserva).exists():
                    Ausencias.objects.create(
                        utilizador=reserva.utilizador,
                        reserva=reserva,
                        taxa_respondida='nao'
                    )

                send_mail(
                    "FITCLUB - Ausência",
                    f"Olá {reserva.utilizador.username},\n\n"
                    f"Fez reserva para o treino do dia {reserva.treino.data_inicio} às {reserva.treino.hora_inicio} "
                    f"mas não compareceu no mesmo.\n"
                    f"Não se esqueça de pagar a taxa de ausência até ao próximo treino!\n\n"
                    "FITCLUB – Mantém-te FIT! Mantém-te Forte!\n\n"
                    "Email gerado automaticamente. Qualquer questão entre em contacto no WhatsApp.",
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[reserva.utilizador.email],
                    fail_silently=False,
                )

        # elif action == 'remover_reserva':
        #     reservas_id = request.POST.get('reservas_id')
        #     reserva = get_object_or_404(Reservas, id=reservas_id)
        #     reserva.delete()
        #     promover_lista_espera(treino)

        elif action == 'remover_reserva':
            reservas_id = request.POST.get('reservas_id')

            with transaction.atomic():
                treino_locked = Treino.objects.select_for_update().get(id=treino.id)

                reserva = Reservas.objects.select_for_update().get(id=reservas_id)
                reserva.delete()

                primeiro = ListaEspera.objects.select_for_update().filter(
                    treino=treino_locked
                ).order_by('data_entrada').first()

                if primeiro:
                    Reservas.objects.create(
                        utilizador=primeiro.utilizador,
                        treino=treino_locked
                    )
                    primeiro.delete()

        elif action == 'remover_espera':
            espera_id = request.POST.get('espera_id')
            espera = get_object_or_404(ListaEspera, id=espera_id)
            espera.delete()

        return redirect('reservas_detalhes', treino_id=treino.id)

    return render(
        request, 
        'FC_RESERVAS/fcReservas_detalhes.html', 
        {'treino': treino, 'reservas': reservas, 'lista_espera': lista_espera}
    )



def promover_lista_espera(treino):
    """Promove o primeiro da lista de espera para a reserva caso haja vaga disponível."""
    primeiro_na_lista = ListaEspera.objects.filter(treino=treino).order_by('data_entrada').first()

    if primeiro_na_lista:
        # Criar a reserva para o primeiro da lista de espera
        nova_reserva = Reservas.objects.create(utilizador=primeiro_na_lista.utilizador, treino=treino)
        
        # Remover da lista de espera
        primeiro_na_lista.delete()

        # Enviar email de notificação
        send_mail(
            subject=f"FITCLUB - Reserva confirmada{treino.data_inicio.strftime('%d/%m/%Y')} às {treino.hora_inicio}",
            message=f"Olá {nova_reserva.utilizador.username},\n\n"
                    f"Alguém cancelou a reserva para o treino, e a tua reserva foi automaticamente confirmada.\n"
                    f"O treino acontecerá no dia {treino.data_inicio.strftime('%d/%m/%Y')} às {treino.hora_inicio}.\n"
                    f"FITCLUB – Mantém-te FIT! Mantém-te Forte!\n"
                    "Email gerado automaticamente. Qualquer questão entre em contacto no WhatsApp.",
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=[nova_reserva.utilizador.email],
            fail_silently=False,
        )





@receiver(post_save, sender=Reservas)
def enviar_email_reserva_lista_espera(sender, instance, created, **kwargs):
    if created:  # Só interessa reservas recém-criadas
        treino = instance.treino

        # Verifica se essa reserva foi feita transferindo alguém da lista de espera
        primeiro_na_lista = ListaEspera.objects.filter(treino=treino).order_by('data_entrada').first()
        if primeiro_na_lista:
            # Obtém a data de início do treino (data_inicio)
            data_treino = treino.data_inicio.strftime('%d/%m/%Y')  # Formato da data: dd/mm/yyyy

            # Envia o e-mail para o usuário promovido
            send_mail(
                subject=f"Reserva confirmada para o treino dia {data_treino} às {treino.hora_inicio}",
                message=(
                    f"Olá {primeiro_na_lista.utilizador.username},\n\n"
                    f"Alguém cancelou a reserva para o treino, e a tua reserva foi automaticamente confirmada.\n"
                    f"O treino acontecerá no dia {data_treino} às {treino.hora_inicio}.\n\n"
                ),
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[primeiro_na_lista.utilizador.email],
                fail_silently=False,
            )





#CRUD

@login_required                                                                    
def editar_treino(request, treino_id):
    if not request.user.is_staff:
        return redirect('calendario')
    
    treino = get_object_or_404(Treino, id=treino_id)

    if request.method == 'POST':
        form = EditarTreinoForm(request.POST, instance=treino)
        if form.is_valid():
            form.save()
            return redirect('calendario')  # Redirecione para a página de lista de treinos
    else:
        form = EditarTreinoForm(instance=treino)

    return render(request, 'CRUD/fcEditar_treino.html', {'form': form, 'treino': treino})


@login_required
def apagartreino(request, pk):
    if not request.user.is_staff:
        return redirect('calendario')

    treino = Treino.objects.filter(pk=pk).first()
    if not treino:
        return redirect('calendario')

    if request.method == 'POST':

        detalhes_treino_cancelado = f"Treino de {treino.dia_da_semana} - {treino.data_inicio} às {treino.hora_inicio} foi cancelado"
        treino.delete()
        utilizadores = Utilizadores.objects.all()
        emails = [utilizador.email for utilizador in utilizadores if utilizador.email]

        send_mail(
            subject="FITCLUB - Treino Cancelado",
            message=f"Olá,\n\nO {detalhes_treino_cancelado}\n\n\nCom os melhores cumprimentos"
            f'FITCLUB - Mantém-te FIT! Mantém-te Forte!\n'
            'Email gerado automaticamente. Qualquer questão entre em contacto no WhatsApp.',

            from_email=settings.EMAIL_HOST_USER,
            recipient_list=emails,
            fail_silently=False,
        )

        return redirect('calendario')

    return render(request, 'CRUD/fcApagar_tarefa.html', {'treino': treino})








@login_required
def apagar_treinos_em_massa(request):
    if not request.user.is_staff:
        return redirect('calendario')

    if request.method == 'POST':
        # Dados do formulário para filtragem
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        tipo_treino = request.POST.get('tipo_treino')
        dia_da_semana = request.POST.get('dia_da_semana')
        hora_inicio = request.POST.get('hora_inicio')

        # Filtro inicial vazio
        filtros = Q()

        # Aplica filtros de acordo com os critérios preenchidos
        if data_inicio:
            filtros &= Q(data_inicio__gte=data_inicio)
        if data_fim:
            filtros &= Q(data_inicio__lte=data_fim)
        if tipo_treino:
            filtros &= Q(tipo_treino=tipo_treino)
        if dia_da_semana:
            filtros &= Q(dia_da_semana=dia_da_semana)
        if hora_inicio:
            filtros &= Q(hora_inicio=hora_inicio)

        # Apaga os treinos filtrados
        treinos_apagados = Treino.objects.filter(filtros).delete()

        # Redireciona após apagar
        return redirect('calendario')

    # Renderiza a página com o formulário
    return render(request, 'CRUD/fcApagar_treinos_em_massa.html')



@login_required
def recordes(request):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    if request.method == 'POST':
        form = CriarRecordesForm(request.POST, user=request.user, utilizador_alvo=request.user)
        if form.is_valid():
            recorde = form.save(commit=False)
            recorde.utilizador = request.user
            recorde.save()

            return redirect('recordes')
    else:
        form = CriarRecordesForm(user=request.user, utilizador_alvo=request.user)

    # Captura o valor do filtro na URL
    query = request.GET.get('q')

    # Filtra os recordes com base no nome do recorde ou no exercício predefinido
    if query:
        recorde = Recordes.objects.filter(
            utilizador=request.user
        ).filter(
            Q(nome__nome__icontains=query) | Q(predefinidos__icontains=query)
        )
    else:
        recorde = Recordes.objects.filter(utilizador=request.user)

    # Obtenha todos os nomes do modelo RecordesNomes
    nomes_disponiveis = RecordesNomes.objects.filter(utilizador=request.user)

    # OBTEM OS EXERCÍCIOS PREDEFINIDOS
    predefinidos_disponiveis = [
        {"value": choice[0], "label": choice[1]}
        for choice in Recordes.EXERCICIOS_PREDEFINIDOS_CHOICES
        if choice[0] != "vazio"  # Exclui a opção vazia
    ]

    # Combine as opções para o filtro
    opcoes_filtro = list(nomes_disponiveis.values_list('nome', flat=True)) + [
        choice["label"] for choice in predefinidos_disponiveis
    ]

    return render(request,'FC_APP/fcCriarRecorde.html',{'form': form,'recorde': recorde,'nomes_disponiveis': nomes_disponiveis,'predefinidos_disponiveis': predefinidos_disponiveis,'opcoes_filtro': opcoes_filtro,'query': query})




@login_required
def apagar_recorde(request, recorde_id):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')

    recorde = get_object_or_404(Recordes, id=recorde_id)
    
    if request.method == 'POST':
        recorde.delete()
        return redirect('recordes')




@login_required
def criarnomerecordes(request):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    if request.method == 'POST':
        form = CriarNomeRecordeForm(request.POST or None)
        if form.is_valid():
            nome = form.save(commit=False)
            nome.utilizador = request.user
            nome.save()

            return redirect('recordes')
    else:
        form = CriarNomeRecordeForm(request.POST)

    nome = RecordesNomes.objects.filter(utilizador=request.user)

    return render(request, 'FC_APP/fcCriar_recorde_nome.html', {'form': form, 'nome': nome})


@login_required
def apagar_nome_recorde(request, recorde_id):
    recorde = get_object_or_404(RecordesNomes, id=recorde_id, utilizador=request.user)

    if request.method == 'POST':
        recorde.delete()
        return redirect('criarnomerecordes')

    return redirect('criarnomerecordes')


@login_required
def ver_recordes_utilizador(request, user_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    utilizador = get_object_or_404(Utilizadores, id=user_id)

    # Formulário de criação de recordes
    if request.method == 'POST':
        form = CriarRecordesForm(request.POST, user=request.user, utilizador_alvo=utilizador)
        if form.is_valid():
            recorde = form.save(commit=False)
            recorde.utilizador = utilizador
            recorde.save()
            return redirect('ver_recordes_utilizador', user_id=utilizador.id)
    else:
        form = CriarRecordesForm(user=request.user, utilizador_alvo=utilizador)

    # Buscar recordes criados pelo utilizador
    recordes_personalizados = Recordes.objects.filter(utilizador=utilizador)

    # Buscar recordes predefinidos
    predefinidos_disponiveis = [
        {"nome": choice[1], "valor": None}
        for choice in Recordes.EXERCICIOS_PREDEFINIDOS_CHOICES
        if choice[0] != "vazio"
    ]

    for predefinido in predefinidos_disponiveis:
        recorde_existente = Recordes.objects.filter(
            utilizador=utilizador,
            predefinidos=predefinido["nome"],
        ).first()
        if recorde_existente:
            predefinido["valor"] = recorde_existente.valor
            predefinido["data_do_recorde"] = recorde_existente.data_do_recorde

    todos_recordes = list(recordes_personalizados) + predefinidos_disponiveis

    return render(request, 'FC_APP/fcVerRecordes.html', {
        'utilizador': utilizador,
        'recordes': todos_recordes,
        'form': form,
    })



@login_required
def assiduidade(request):
    if request.user.funcao != 'Ativo':
        return redirect('acesso_negado')
    
    # Obtém o mês e o ano da URL ou usa o atual
    ano = int(request.GET.get('ano', datetime.now().year))
    mes = int(request.GET.get('mes', datetime.now().month))

    # Busca as reservas do usuário para o mês e ano selecionados
    reservas = Reservas.objects.filter(
        utilizador=request.user,
        treino__data_inicio__year=ano,
        treino__data_inicio__month=mes,
    ).select_related('treino')

    # Contar treinos confirmados
    total_reservas_confirmadas = reservas.filter(confirmado=True).count()

    # Contar treinos não confirmados
    total_reservas_nao_confirmadas = reservas.filter(confirmado=False).count()

    # Contar o total de treinos
    total_reservas = reservas.count()

    # Calcular a média semanal de frequência (5 dias úteis por semana)
    _, total_dias_no_mes = monthrange(ano, mes)
    semanas_uteis_no_mes = ceil(total_dias_no_mes / 7)  # Total de semanas no mês
    dias_uteis_totais = semanas_uteis_no_mes * 5  # Máximo de treinos possíveis no mês

    # Calcular a média semanal como porcentagem
    if dias_uteis_totais > 0:
        media_semanal = (total_reservas_confirmadas / dias_uteis_totais) * 100
    else:
        media_semanal = 0

    # Filtro por meses
    meses = range(1, 13)

    return render(
        request,
        'FC_APP/fcAssiduidade.html',
        {
            'ano': ano,
            'mes': mes,
            'meses': meses,
            'total_reservas_confirmadas': total_reservas_confirmadas,
            'total_reservas_nao_confirmadas': total_reservas_nao_confirmadas,
            'total_reservas': total_reservas,
            'media_semanal': media_semanal,
            'reservas': reservas,
        }
    )


@login_required
def ver_assiduidade(request, user_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    # Obtém o mês e ano da query string ou usa os valores atuais
    ano = int(request.GET.get('ano', datetime.now().year))
    mes = int(request.GET.get('mes', datetime.now().month))

    # Busca o utilizador para mostrar no template
    utilizador = Utilizadores.objects.get(id=user_id)

    # Filtra as reservas com base no mês, ano e utilizador
    reservas = Reservas.objects.filter(
        utilizador_id=user_id,
        treino__data_inicio__year=ano,
        treino__data_inicio__month=mes,
    ).select_related('treino')

    # Conta o número de reservas confirmadas
    total_reservas_confirmadas = reservas.filter(confirmado=True).count()

    # Conta o número de reservas não confirmadas
    total_reservas_nao_confirmadas = reservas.filter(confirmado=False).count()

    # Conta o número total de reservas
    total_reservas = reservas.count()

    # Cálculo da média semanal de frequência
    # Número de dias no mês
    _, total_dias_no_mes = monthrange(ano, mes)
    # Número de semanas úteis no mês (5 dias úteis por semana)
    semanas_uteis_no_mes = ceil(total_dias_no_mes / 7)
    dias_uteis_totais = semanas_uteis_no_mes * 5  # Máximo de dias úteis no mês

    # Calcular a média semanal como porcentagem
    if dias_uteis_totais > 0:
        media_semanal = (total_reservas_confirmadas / dias_uteis_totais) * 100
    else:
        media_semanal = 0

    # Lista de meses para o dropdown
    meses = range(1, 13)

    return render(
        request,
        'FC_APP/fcVer_assiduidade.html',  # Template adaptado
        {
            'utilizador': utilizador,
            'reservas': reservas,
            'ano': ano,
            'mes': mes,
            'meses': meses,
            'total_reservas': total_reservas,
            'total_reservas_confirmadas': total_reservas_confirmadas,
            'total_reservas_nao_confirmadas': total_reservas_nao_confirmadas,
            'media_semanal': media_semanal,
        },
    )



from openpyxl import Workbook


@login_required
def exportar_assiduidade_estatisticas_excel(request):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    # Obter o mês e o ano do input
    mes_input = request.GET.get('mes')  # tipo '2025-10'
    if mes_input and '-' in mes_input:
        ano_str, mes_str = mes_input.split('-')
        ano = int(ano_str)
        mes = int(mes_str)
    else:
        ano = datetime.now().year
        mes = datetime.now().month

    # Criar o Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Assiduidade_{mes:02d}-{ano}"

    # Cabeçalhos
    ws.append(["Utilizador", "Total de Treinos", "Confirmados", "Não Confirmados", "Média Semanal (%)"])

    # Iterar por todos os utilizadores
    utilizadores = Utilizadores.objects.all().order_by('username')

    for u in utilizadores:
        # Filtrar reservas do utilizador nesse mês
        reservas = Reservas.objects.filter(
            utilizador=u,
            treino__data_inicio__year=ano,
            treino__data_inicio__month=mes,
        )

        total_reservas = reservas.count()
        total_confirmadas = reservas.filter(confirmado=True).count()
        total_nao_confirmadas = reservas.filter(confirmado=False).count()

        # Cálculo da média semanal
        _, total_dias_no_mes = monthrange(ano, mes)
        semanas_uteis_no_mes = ceil(total_dias_no_mes / 7)
        dias_uteis_totais = semanas_uteis_no_mes * 5  # considerando 5 dias úteis por semana

        if dias_uteis_totais > 0:
            media_semanal = (total_confirmadas / dias_uteis_totais) * 100
        else:
            media_semanal = 0

        # Adicionar linha
        ws.append([
            u.username,
            total_reservas,
            total_confirmadas,
            total_nao_confirmadas,
            round(media_semanal, 2)
        ])

    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"assiduidade_{ano}-{mes:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response



#TESTES

# VER SE EXPORTA OS DADOS DE TODOS OS UTILIZADORES
@login_required
def export_to_excel(request):
    # Cria uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados Biométricos"

    # Cabeçalho da planilha
    columns = ['ID', 'Nome', 'Email', 'Contacto', 'Data de Nascimento', 'Morada', 'Código Postal', 'Localidade',
                'NIF', 'Pretende Recibo', 'Profissão', 'Classificação de Esforço na Profissão', 'Fumador', 'Problemas de Saúde', 'Limitações para a prática de exercício Físico']
    ws.append(columns)

    # VAI BUSCAR OS DADOS A BASE DE DADOS
    informacoes = Utilizadores.objects.all()

    # DADOS QUE VÃO SER PASSADOS PARA O EXCEL
    for data in informacoes:
        row = [
            data.id, 
            data.username, 
            data.email, 
            data.contacto, 
            data.data_nascimento, 
            data.morada, 
            data.codigo_postal, 
            data.localidade,
            data.nif,
            data.pretende_recibo,
            data.profissao,
            data.classificacao_esforco_na_profissao,
            data.fumador,
            data.problemas_saude,
            data.limitacoes_para_pratica_exercicio_fisico,
        ]  # Altere conforme os campos do seu modelo
        ws.append(row)

    # Criar o response com o arquivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="dados_totais.xlsx"'

    # Salvar o arquivo na resposta HTTP
    wb.save(response)
    return response





# EXPORTA OS DADOS DE UM UTILIZADOR ESPECIFICO
@login_required
def export_user_data_to_excel(request, user_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    # BUSCAR O UTILIZADOR ESPECIFICO PELO ID
    utilizador = get_object_or_404(Utilizadores, id=user_id)

    # Criar uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Dados do utilizador {utilizador.username}"

    # Cabeçalho da planilha
    columns = ['ID', 'Nome', 'Email', 'Contacto', 'Data de Nascimento', 'Morada', 'Código Postal', 'Localidade',
                'NIF', 'Pretende Recibo', 'Profissão', 'Classificação de Esforço na Profissão', 'Fumador', 'Problemas de Saúde', 'Limitações para a prática de exercício Físico']
    ws.append(columns)

    # Adicionar os dados do utilizador selecionado
    row = [
        utilizador.id, 
        utilizador.username, 
        utilizador.email, 
        utilizador.contacto, 
        utilizador.data_nascimento, 
        utilizador.morada, 
        utilizador.codigo_postal, 
        utilizador.localidade,
        utilizador.nif,
        utilizador.pretende_recibo,
        utilizador.profissao,
        utilizador.classificacao_esforco_na_profissao,
        utilizador.fumador,
        utilizador.problemas_saude,
        utilizador.limitacoes_para_pratica_exercicio_fisico,
    ]
    ws.append(row)

    # Criar a resposta com o arquivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="dados_utilizador_{utilizador.username}.xlsx"'

    # Salvar o arquivo na resposta HTTP
    wb.save(response)
    return response

import csv

@login_required
def baixar_avaliacoes_mensais(request):
    ano = int(request.GET.get('ano', date.today().year))
    mes = int(request.GET.get('mes', date.today().month - 1 or 12))

    avaliacoes_mensais = AvaliacaoMensal.objects.filter(
        ano=ano,
        mes=mes
    ).select_related('utilizador')

    avaliacoes_trimestrais = AvaliacaoTrimestral.objects.filter(
        ano=ano,
        mes=mes
    ).select_related('utilizador')

    wb = Workbook()
    ws = wb.active
    ws.title = f"Avaliacoes_{mes}_{ano}"

    ws.append(['Utilizador', 'Mes', 'Ano', 'Nível Satisfação - Mensal', 'Nível Satisfação - Trimestral'])

    for a in avaliacoes_mensais:
        ws.append([
            a.utilizador.username,
            a.mes,
            str(a.ano),
            a.nivel_satisfacao
        ])

    for a in avaliacoes_trimestrais:
        ws.append([
            a.utilizador.username,
            a.mes,
            str(a.ano),
            '-----',
            a.nivel_satisfacao,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="avaliacoes_mensais_{mes}_{ano}.xlsx"'
        
    wb.save(response)
    return response


from django.db.models import Count


@login_required
def exportar_treinos(request):
    # Datas obrigatórias
    data_inicio_str = request.GET.get('data_inicio')
    data_fim_str = request.GET.get('data_fim')

    if not data_inicio_str or not data_fim_str:
        return HttpResponse("Por favor, forneça data de início e data de fim.", status=400)

    try:
        data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d").date()
        data_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Formato de data inválido. Use YYYY-MM-DD.", status=400)

    treinos = Treino.objects.filter(
        data_inicio__gte=data_inicio,
        data_fim__lte=data_fim
    ).select_related('tipo_treino_nome').annotate(numero_pessoas=Count('reservas')).order_by('data_inicio', 'hora_inicio')

    wb = Workbook()
    ws = wb.active
    ws.title = f"Treinos_{data_inicio}_{data_fim}"

    ws.append(['Dia', 'Tipo Treino', 'Tipo Treino Personalizado', 'Hora Início', 'Hora Fim', 'Reservas', 'Máximo Participantes'])

    for t in treinos:
        ws.append([
            t.data_inicio.strftime("%Y-%m-%d"),
            t.get_tipo_treino_display(),
            getattr(t.tipo_treino_nome, 'nome', ''),
            t.hora_inicio.strftime("%H:%M"),
            t.hora_fim.strftime("%H:%M"),
            t.numero_pessoas,
            t.max_participantes,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="treinos_{data_inicio}_{data_fim}.xlsx"'

    wb.save(response)
    return response





# def exportar_assiduidade_excel(request):
#     # --- 1. Obter o mês a partir do query param ---
#     mes = request.GET.get('mes')  # formato: '2025-10'
#     if not mes:
#         return HttpResponse("Por favor, especifique o mês (?mes=YYYY-MM).")

#     ano, mes_num = mes.split('-')
#     ano, mes_num = int(ano), int(mes_num)

#     # --- 2. Filtrar os registos da assiduidade ---
#     assiduidades = Assiduidade.objects.filter(
#         data__year=ano,
#         data__month=mes_num,
#         utilizador=request.user
#     )

#     # --- 3. Criar o workbook ---
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Assiduidade"

#     # --- 4. Cabeçalhos ---
#     ws.append(["Data", "Presença", "Hora Início", "Hora Fim", "Duração (min)"])

#     # --- 5. Preencher linhas ---
#     for a in assiduidades:
#         ws.append([
#             a.data.strftime("%d/%m/%Y"),
#             "Presente" if a.presente else "Falta",
#             a.hora_inicio.strftime("%H:%M") if a.hora_inicio else "",
#             a.hora_fim.strftime("%H:%M") if a.hora_fim else "",
#             a.duracao_minutos() if hasattr(a, 'duracao_minutos') else "",
#         ])

#     # --- 6. Gerar resposta HTTP ---
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     filename = f"assiduidade_{mes}.xlsx"
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'

#     wb.save(response)
#     return response





from django.contrib.auth.tokens import default_token_generator

def password_reset_request(request):
    if request.method == "POST":
        form = PasswordResetRequestForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]
            user = Utilizadores.objects.get(email=email)

            # Criando o token e a URL de recuperação
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            
            current_site = request.get_host()
            reset_url = f"http://{current_site}/reset/{uid}/{token}/"

            # Enviar o e-mail
            send_mail(
                f'FITCLUB - Redefinir Password',
                f'Olá {user.username},\n\n'
                f'Clica no link para redefinir a tua password: {reset_url}\n\n'
                f'FITCLUB – Mantém-te FIT! Mantém-te Forte!\n'
                f'Email gerado automaticamente. Qualquer questão entre em contacto no WhatsApp.',
                settings.EMAIL_HOST_USER,
                [email],
                fail_silently=False,
            )

            return redirect("login")
    else:
        form = PasswordResetRequestForm()

    return render(request, "CONTAS/password_reset_form.html", {"form": form})





def password_reset_confirm(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = Utilizadores.objects.get(pk=uid)
    except (Utilizadores.DoesNotExist, ValueError, TypeError):
        user = None

    if user and default_token_generator.check_token(user, token):
        if request.method == "POST":
            form = SetNewPasswordForm(request.POST)
            if form.is_valid():
                user.set_password(form.cleaned_data["new_password1"])
                user.save()
                return redirect('login')
        else:
            form = SetNewPasswordForm()
        return render(request, "CONTAS/password_reset_confirm.html", {"form": form})
    else:
        return render(request, "CONTAS/password_reset_invalid.html")


@login_required
def cartao(request):
    return render(request, 'FC_APP/fcCartao.html')



@login_required
def criar_treino_online(request):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')
    
    if request.method == 'POST':
        form = CriarAulaOnline(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('fcBase')
    else:
        form = CriarAulaOnline()

    return render(request, 'TreinosOnline/criar_aula.html', {'form': form})



@login_required
def treinos_funcionais(request):
    funcionais = TreinosOnline.objects.filter(tipo_treino='funcional')
    return render(request, 'TreinosOnline/funcional.html', {'funcionais': funcionais})


@login_required
def apagar_funcionais_online(request, treino_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    treino = get_object_or_404(TreinosOnline, id=treino_id)
    treino.delete()
    return redirect('funcionais')



@login_required
def treinos_mobilidade(request):
    mobilidade = TreinosOnline.objects.filter(tipo_treino='mobilidade')
    return render(request, 'TreinosOnline/mobilidade.html', {'mobilidade': mobilidade})


@login_required
def apagar_mobilidade_online(request, treino_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    treino = get_object_or_404(TreinosOnline, id=treino_id)
    treino.delete()
    return redirect('mobilidade')



@login_required
def treinos_forca(request):
    forca = TreinosOnline.objects.filter(tipo_treino='força')
    return render(request, 'TreinosOnline/força.html', {'forca': forca})



@login_required
def apagar_forca_online(request, treino_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    treino = get_object_or_404(TreinosOnline, id=treino_id)
    treino.delete()
    return redirect('força')



@login_required
def treinos_metabolico(request):
    metabolico = TreinosOnline.objects.filter(tipo_treino='metabólico')
    return render(request, 'TreinosOnline/metabolico.html', {'metabolico': metabolico})



@login_required
def apagar_metabolico_online(request, treino_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    treino = get_object_or_404(TreinosOnline, id=treino_id)
    treino.delete()
    return redirect('metabolico')


@login_required
def treinos_aerobica(request):
    aerobica = TreinosOnline.objects.filter(tipo_treino='aeróbica')
    return render(request, 'TreinosOnline/aerobica.html', {'aerobica': aerobica})


@login_required
def apagar_aerobica_online(request, treino_id):
    if not request.user.is_staff:
        return render(request, 'ERRORS/403.html')

    treino = get_object_or_404(TreinosOnline, id=treino_id)
    treino.delete()
    return redirect('aerobica')



@login_required
def contactos_direitos_imagem(request):
    return render(request, 'REGULAMENTOS/contactos_direitos_imagem.html')


@login_required
def marcacoes_cancelamentos(request):
    return render(request, 'REGULAMENTOS/marcacoes_cancelamentos.html')


@login_required
def mensalidades(request):
    return render(request, 'REGULAMENTOS/mensalidades.html')


@login_required
def regras_treino(request):
    return render(request, 'REGULAMENTOS/regras_treino.html')


@login_required
def suspensao_cancelamento_inscricao(request):
    return render(request, 'REGULAMENTOS/suspencao_cancelamento_inscricao.html')


@login_required
def descarregar_dados(request):
    ano_atual = date.today().year
    anos = list(range(ano_atual - 2, ano_atual + 1))
    meses = list(range(1, 13))

    return render(request, 'descarregar_dados.html', {'anos': anos, 'meses': meses})

